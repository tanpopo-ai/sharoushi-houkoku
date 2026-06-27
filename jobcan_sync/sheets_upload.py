"""
ジョブカンからダウンロードしたCSVをGoogleスプレッドシートに追記する。

OAuth認証方式: 初回実行時にブラウザが開き、Googleアカウントでスクリプトに
スプレッドシート編集を許可する。許可情報は token.json として保存され、
2回目以降は自動更新される (タスクスケジューラからも動作する)。

シート末尾にデータを append するシンプル方式。最初の取得時はCSVヘッダ +
"取得日時" 列を入れて、以降の取得では取得日時を付けてデータ行のみを追記する。
"""
from __future__ import annotations

import csv
import json
import logging
import os
import time
from datetime import datetime
from pathlib import Path

import gspread
import gspread.exceptions
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow

logger = logging.getLogger(__name__)


def _col_letter(idx: int) -> str:
    """0-indexed の列インデックス → 'A','B',...,'Z','AA',... の列文字。"""
    s = ""
    n = idx
    while True:
        s = chr(ord('A') + n % 26) + s
        n = n // 26 - 1
        if n < 0:
            break
    return s


def _api_retry(func, *args, _max_retries: int = 8, _initial_delay: int = 4, **kwargs):
    """gspread API 呼び出しを 429(rate limit) / 5xx に対してリトライする。"""
    delay = _initial_delay
    last_err: Exception | None = None
    for attempt in range(_max_retries):
        try:
            return func(*args, **kwargs)
        except gspread.exceptions.APIError as e:
            last_err = e
            status = None
            try:
                status = e.response.status_code
            except Exception:
                pass
            if status in (429, 500, 502, 503, 504):
                wait = min(delay, 60)
                logger.warning(
                    "Sheets API status=%s 検出。%d秒待機後リトライ (試行 %d/%d)",
                    status, wait, attempt + 1, _max_retries,
                )
                time.sleep(wait)
                delay = min(delay * 2, 60)
            else:
                raise
    if last_err:
        raise last_err
    raise RuntimeError("API リトライ上限到達")

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]
ACQUIRED_AT_HEADER = "取得日時"


def _load_oauth_credentials(config: dict):
    """Google認証情報を読み込む。
    クラウド(GitHub Actions)では環境変数 GOOGLE_SERVICE_ACCOUNT_JSON の
    サービスアカウントを優先。無ければ従来のOAuth(ローカルPC用)。"""
    sa_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if sa_json:
        from google.oauth2.service_account import Credentials as ServiceAccountCredentials
        info = json.loads(sa_json)
        logger.info("サービスアカウント認証を使用します (client_email=%s)",
                    info.get("client_email", "?"))
        return ServiceAccountCredentials.from_service_account_info(
            info,
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
            ],
        )

    g = config["google"]
    client_secret_path = Path(g["oauth_client_secret"]).resolve()
    token_path = Path(g.get("oauth_token", "token.json")).resolve()

    creds: Credentials | None = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)

    if creds and creds.valid:
        return creds

    if creds and creds.expired and creds.refresh_token:
        logger.info("アクセストークンを更新します。")
        creds.refresh(Request())
        token_path.write_text(creds.to_json(), encoding="utf-8")
        return creds

    if not client_secret_path.exists():
        raise FileNotFoundError(
            f"OAuthクライアント機密ファイルが見つかりません: {client_secret_path}. README参照。"
        )

    logger.info("初回認可: ブラウザを開いてGoogleアカウントで許可してください。")
    flow = InstalledAppFlow.from_client_secrets_file(str(client_secret_path), SCOPES)
    creds = flow.run_local_server(port=0)
    token_path.write_text(creds.to_json(), encoding="utf-8")
    logger.info("認可情報を %s に保存しました。", token_path)
    return creds


def _open_worksheet(config: dict) -> gspread.Worksheet:
    g = config["google"]
    creds = _load_oauth_credentials(config)
    client = gspread.authorize(creds)
    sheet = client.open_by_key(g["spreadsheet_id"])

    try:
        ws = sheet.worksheet(g["worksheet_name"])
    except gspread.WorksheetNotFound:
        logger.info("ワークシート '%s' が存在しないので作成します。", g["worksheet_name"])
        ws = sheet.add_worksheet(title=g["worksheet_name"], rows=1000, cols=40)
    return ws


def _read_csv(csv_path: Path) -> tuple[list[str], list[list[str]]]:
    # ジョブカンCSVは Shift_JIS で出力されることが多い。失敗時は UTF-8 にフォールバック。
    for enc in ("cp932", "utf-8-sig", "utf-8"):
        try:
            with csv_path.open("r", encoding=enc, newline="") as f:
                rows = list(csv.reader(f))
            logger.info("CSVを %s で読み込みました (%d行)", enc, len(rows))
            break
        except UnicodeDecodeError:
            continue
    else:
        raise RuntimeError(f"CSVのエンコーディング判定に失敗: {csv_path}")

    if not rows:
        return [], []
    return rows[0], rows[1:]


def _safe_append(ws, rows: list[list], batch_size: int = 200) -> int:
    """ws の最終行の次に rows を書き込む。
    append_rows のテーブル境界自動推定で列ズレが起きないよう、明示的な範囲指定で書く。
    シートの行数が足りないと API エラーになるため、必要に応じて自動拡張する。
    """
    if not rows:
        return 0
    existing = ws.get_all_values()
    next_row = len(existing) + 1

    # シート行数の事前拡張 (margin 200)
    needed_rows = next_row + len(rows) + 200
    if ws.row_count < needed_rows:
        try:
            ws.add_rows(needed_rows - ws.row_count)
            logger.info("ws.add_rows: %d → %d", ws.row_count, needed_rows)
        except Exception as e:
            logger.warning("シート行数拡張失敗: %s", e)

    total = 0
    for i in range(0, len(rows), batch_size):
        chunk = rows[i:i + batch_size]
        start_row = next_row + i
        range_name = f"A{start_row}"
        ws.update(range_name=range_name, values=chunk, value_input_option="USER_ENTERED")
        total += len(chunk)
    return total


def upload_csv(config: dict, csv_path: Path) -> int:
    """CSVをスプレッドシートに追記し、追加した行数を返す。"""
    ws = _open_worksheet(config)
    header, data_rows = _read_csv(csv_path)

    if not data_rows:
        logger.warning("CSVにデータ行がありません: %s", csv_path)
        return 0

    acquired_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    existing = ws.row_values(1)
    if not existing:
        ws.update(range_name="A1", values=[[ACQUIRED_AT_HEADER, *header]],
                  value_input_option="USER_ENTERED")
        logger.info("ヘッダ行を書き込みました(%d列)", len([ACQUIRED_AT_HEADER, *header]))

    payload = [[acquired_at, *r] for r in data_rows]
    n = _safe_append(ws, payload)
    logger.info("%d行を追記しました。", n)
    return n


def upload_attendance_book(config: dict, file_path: Path) -> int:
    """出勤簿一括ダウンロードのファイルを読み、全スタッフ分をフラットに
    1つのワークシートに追記する。

    入力ファイルは以下のいずれか:
      - .zip  : スタッフごとに個別の .xlsx が格納
      - .xlsx : 1ファイル内の複数シートにスタッフ別データ

    各シート内で "日付" を含むヘッダ行を探し、その下を勤怠データとみなす。
    """
    g = config["google"]
    ws_name = g.get("attendance_book_worksheet_name", "出勤簿ログ")

    # ファイルをパースして (スタッフラベル, 行配列) のリストに変換
    sources = _load_attendance_sources(file_path)
    logger.info("出勤簿: %d スタッフ分のシートを取得", len(sources))

    acquired_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_rows: list[list[str]] = []
    canonical_header: list[str] | None = None

    for staff_label, rows in sources:
        header_idx = _find_header_row(rows)
        if header_idx is None:
            logger.warning("スタッフ '%s': ヘッダ行(日付列)が見つかりません。スキップ。", staff_label)
            continue

        sheet_header = [_norm(c) for c in rows[header_idx]]
        if canonical_header is None:
            canonical_header = [ACQUIRED_AT_HEADER, "スタッフ", *sheet_header]

        for r in rows[header_idx + 1:]:
            if all(c is None or _norm(c) == "" for c in r):
                break  # 空行で終了
            all_rows.append([acquired_at, staff_label, *[_norm(c) for c in r]])

    if not all_rows:
        logger.warning("出勤簿に有効データなし: %s", file_path)
        return 0

    # ワークシート取得・作成
    creds = _load_oauth_credentials(config)
    client = gspread.authorize(creds)
    sheet = client.open_by_key(g["spreadsheet_id"])
    try:
        ws = sheet.worksheet(ws_name)
    except gspread.WorksheetNotFound:
        logger.info("ワークシート '%s' を作成", ws_name)
        ws = sheet.add_worksheet(title=ws_name, rows=5000, cols=40)

    existing = ws.row_values(1)
    if not existing and canonical_header:
        ws.update(range_name="A1", values=[canonical_header],
                  value_input_option="USER_ENTERED")
        logger.info("出勤簿ログ ヘッダ書き込み: %d列", len(canonical_header))

    # 明示的なセル範囲指定でバッチ書き込み
    BATCH = 200
    total = 0
    base_existing = ws.get_all_values()
    next_row = len(base_existing) + 1
    # シート行数の事前拡張
    needed_rows = next_row + len(all_rows) + 500
    if ws.row_count < needed_rows:
        try:
            ws.add_rows(needed_rows - ws.row_count)
            logger.info("出勤簿ログ: 行数を %d に拡張", needed_rows)
        except Exception as e:
            logger.warning("出勤簿ログ 行数拡張失敗: %s", e)
    for i in range(0, len(all_rows), BATCH):
        chunk = all_rows[i:i + BATCH]
        start_row = next_row + i
        ws.update(range_name=f"A{start_row}", values=chunk,
                  value_input_option="USER_ENTERED")
        total += len(chunk)
        logger.info("出勤簿: %d/%d 行追記 (A%d〜)", total, len(all_rows), start_row)

    # スタッフごとに分けたタブも更新
    pss = config.get("per_staff_settings", {})
    if pss.get("enabled", True) and canonical_header:
        try:
            _split_to_per_staff_sheets(config, all_rows, canonical_header, sheet)
        except Exception as e:
            logger.error("スタッフ別シート更新失敗: %s", e, exc_info=True)
        try:
            _create_summary_sheet(config, sheet, ws)
        except Exception as e:
            logger.error("集計タブ更新失敗: %s", e, exc_info=True)

    return total


def _safe_sheet_title(label: str) -> str:
    """Sheets で使えない文字 ( : \\ / ? * [ ] ) を除去し、100文字以内に切り詰める。"""
    bad = ':\\/?*[]'
    clean = "".join(c for c in label if c not in bad).strip()
    return (clean or "(unknown)")[:99]


def _split_to_per_staff_sheets(config: dict, all_rows: list[list[str]],
                                canonical_header: list[str], sh) -> None:
    """all_rows をスタッフ別タブに振り分けて書き込み。数式列(勤務地判定/残業/特出)も付与。

    all_rows は [acquired_at, staff_label, *cells] 形式。
    canonical_header は [取得日時, スタッフ, *xlsx_header] 形式。
    """
    from collections import defaultdict

    pss = config.get("per_staff_settings", {})
    overrides = pss.get("fixed_location_overrides", {})
    th = pss.get("thresholds", {})
    branch_in = th.get("branch_clock_in_at_or_before", "08:30")
    branch_out = th.get("branch_clock_out_at_or_before", "19:00")
    special_th = th.get("special_overtime_threshold", "03:00")

    # スタッフ別にグループ化
    groups: dict[str, list[list[str]]] = defaultdict(list)
    for row in all_rows:
        if len(row) < 2:
            continue
        staff_label = row[1]
        per_staff_row = [row[0]] + list(row[2:])  # acquired_at + xlsx_cells (スタッフ列を除く)
        groups[staff_label].append(per_staff_row)

    base_header = [canonical_header[0]] + list(canonical_header[2:])  # "スタッフ"除外
    full_header = base_header + ["判定勤務地", "残業時間", "特出時間"]

    # 列位置を base_header から動的に検出 (堅牢に)
    def _find_col(name_substr: str) -> str:
        for i, h in enumerate(base_header):
            if name_substr in (h or ""):
                return _col_letter(i)
        return _col_letter(0)

    col_in = _find_col("出勤")          # 出勤時刻
    col_out = _find_col("退勤")          # 退勤時刻
    col_overtime = _find_col("実残業")    # 実残業時間 (残業/特出時間の元データ)
    logger.info("スタッフ別シート: 出勤=%s 退勤=%s 実残業=%s", col_in, col_out, col_overtime)

    total_staff = 0
    total_to_process = len(groups)
    for staff_idx, (staff_label, rows) in enumerate(groups.items(), 1):
      try:
        staff_name = staff_label.split("(")[0].strip()
        fixed_loc = overrides.get(staff_name)

        title = _safe_sheet_title(staff_label)
        try:
            ws = _api_retry(sh.worksheet, title)
        except gspread.WorksheetNotFound:
            ws = _api_retry(
                sh.add_worksheet, title=title,
                rows=2000, cols=max(30, len(full_header)),
            )
            _api_retry(
                ws.update, range_name="A1",
                values=[full_header], value_input_option="USER_ENTERED",
            )
            logger.info("新規スタッフタブ作成 (%d/%d): %s", staff_idx, total_to_process, title)

        existing = _api_retry(ws.get_all_values)
        next_row = max(2, len(existing) + 1)

        # 各行に数式列を付与
        # ヘルパー: セル値を数値時刻に変換する式の断片 (数値でも文字列でも両対応)
        def _t(letter: str, rn: int) -> str:
            return f"IFERROR(TIMEVALUE({letter}{rn}),N({letter}{rn}))"

        over_col_letter = _col_letter(len(base_header) + 1)  # 残業時間
        spec_col_letter = _col_letter(len(base_header) + 2)  # 特出時間
        batch_first_row = next_row  # このバッチで書き込む最初の行(=日次行の開始)

        rows_with_fml: list[list[str]] = []
        for i, row in enumerate(rows):
            r_num = next_row + i
            full_row = list(row)
            while len(full_row) < len(base_header):
                full_row.append("")

            # 日付列(per_staff_row[1]) が "合計" なら合計行と判定
            is_summary = (len(row) > 1 and "合計" in str(row[1] or ""))

            if is_summary:
                # 合計行: R/S は列の合計、Q は空 (判定不要)
                last_data_row = r_num - 1
                if last_data_row >= batch_first_row:
                    fml_over = f'=SUM({over_col_letter}{batch_first_row}:{over_col_letter}{last_data_row})'
                    fml_spec = f'=SUM({spec_col_letter}{batch_first_row}:{spec_col_letter}{last_data_row})'
                else:
                    fml_over = "0"
                    fml_spec = "0"
                fml_loc = ""
            else:
                in_t = _t(col_in, r_num)
                out_t = _t(col_out, r_num)
                ovt_t = _t(col_overtime, r_num)  # 実残業時間
                # 判定勤務地: 出勤時刻=0 なら休み。
                # 分院 = (0 < 出勤時刻 <= branch_in) AND (0 < 退勤時刻 <= branch_out)
                if fixed_loc:
                    fml_loc = f'=IF({in_t}=0,"休み","{fixed_loc}")'
                else:
                    fml_loc = (
                        f'=IF({in_t}=0,"休み",'
                        f'IF(AND({in_t}<=TIMEVALUE("{branch_in}"),'
                        f'{out_t}>0,{out_t}<=TIMEVALUE("{branch_out}")),"分院","本院"))'
                    )
                # 残業 / 特出 (実残業時間を元に分岐)
                fml_over = f'=IF({ovt_t}>=TIMEVALUE("{special_th}"),0,{ovt_t})'
                fml_spec = f'=IF({ovt_t}>=TIMEVALUE("{special_th}"),{ovt_t},0)'

            full_row.extend([fml_loc, fml_over, fml_spec])
            rows_with_fml.append(full_row)

        if rows_with_fml:
            BATCH = 200
            for i in range(0, len(rows_with_fml), BATCH):
                chunk = rows_with_fml[i:i + BATCH]
                start = next_row + i
                _api_retry(
                    ws.update, range_name=f"A{start}",
                    values=chunk, value_input_option="USER_ENTERED",
                )
            logger.info("スタッフタブ (%d/%d) '%s': %d行追記",
                        staff_idx, total_to_process, title, len(rows_with_fml))

            # 残業時間 / 特出時間 の列を時刻フォーマットに (列全体)
            over_col = _col_letter(len(base_header) + 1)
            spec_col = _col_letter(len(base_header) + 2)
            try:
                _api_retry(
                    ws.format,
                    f"{over_col}:{spec_col}",
                    {"numberFormat": {"type": "TIME", "pattern": "[h]:mm"}},
                )
            except Exception as e:
                logger.warning("'%s' 時刻フォーマット適用失敗: %s", title, e)
            total_staff += 1

        # スロットリング: レート制限回避のため小休止
        time.sleep(0.6)
      except Exception as e:
          logger.error("スタッフ '%s' 処理失敗 (続行): %s", staff_label, e, exc_info=True)
          # 諦めずに次のスタッフへ
          continue

    logger.info("スタッフ別シート更新完了: %d / %d タブ", total_staff, total_to_process)


def _parse_time_to_seconds(v) -> int:
    """'1:30' / '01:30:00' / 数値(時刻シリアル) → 秒数。空・不正は0。"""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(round(float(v) * 86400))
    s = str(v).strip()
    if not s:
        return 0
    if ":" in s:
        try:
            parts = s.split(":")
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            sec = int(parts[2]) if len(parts) > 2 else 0
            return h * 3600 + m * 60 + sec
        except ValueError:
            pass
    try:
        return int(round(float(s) * 86400))
    except ValueError:
        return 0


def _create_summary_sheet(config: dict, sh, master_ws) -> None:
    """マスター 出勤簿ログ から全データを読み、スタッフ別×月別の残業/特出合計と
    累積をまとめた「集計」タブを作成し、各スタッフ用の棒グラフを配置する。
    """
    import re
    from collections import defaultdict

    summary_name = config.get("google", {}).get("summary_worksheet_name", "集計")
    pss = config.get("per_staff_settings", {})
    threshold_sec = _parse_time_to_seconds(
        pss.get("thresholds", {}).get("special_overtime_threshold", "03:00")
    )

    all_data = master_ws.get_all_values()
    if len(all_data) < 2:
        logger.warning("集計タブ: 出勤簿ログにデータなし、スキップ")
        return

    header, rows = all_data[0], all_data[1:]

    def _find_idx(needle: str) -> int:
        return next((i for i, h in enumerate(header) if needle in (h or "")), -1)

    idx_acq = _find_idx("取得日時")
    idx_staff = _find_idx("スタッフ")
    idx_date = _find_idx("日付")
    idx_actual = _find_idx("実残業")
    if min(idx_acq, idx_staff, idx_date, idx_actual) < 0:
        logger.warning("集計: ヘッダから必要列が見つからず。header=%s", header)
        return

    # 各 (staff, date) について 取得日時 が最新の行のみ採用 (重複取得を排除)
    latest: dict[tuple[str, str], list[str]] = {}
    max_idx = max(idx_acq, idx_staff, idx_date, idx_actual)
    for row in rows:
        if len(row) <= max_idx:
            continue
        staff = row[idx_staff]
        date_str = row[idx_date]
        if not staff or not date_str or "合計" in str(date_str):
            continue
        key = (staff, date_str)
        if key not in latest or latest[key][idx_acq] < row[idx_acq]:
            latest[key] = row

    # 月別集計
    monthly: dict[tuple[str, int, int], dict[str, int]] = defaultdict(
        lambda: {"over": 0, "spec": 0}
    )
    for (staff, date_str), row in latest.items():
        m = re.match(r"(\d+)/", str(date_str))
        if not m:
            continue
        month_num = int(m.group(1))
        try:
            year = int(str(row[idx_acq])[:4])
        except (ValueError, IndexError):
            continue
        sec = _parse_time_to_seconds(row[idx_actual])
        if sec >= threshold_sec:
            monthly[(staff, year, month_num)]["spec"] += sec
        else:
            monthly[(staff, year, month_num)]["over"] += sec

    if not monthly:
        logger.warning("集計タブ: データなし")
        return

    # 月とスタッフの一覧 (集計テーブルの列・行)
    all_months: list[tuple[int, int]] = sorted({(y, m) for (_, y, m) in monthly})
    all_staff: list[str] = sorted({s for (s, _, _) in monthly})

    # 集計タブを再作成
    try:
        old_ws = _api_retry(sh.worksheet, summary_name)
        _api_retry(sh.del_worksheet, old_ws)
    except gspread.WorksheetNotFound:
        pass
    needed_rows = max(2000, (len(all_staff) + 35) * 4 + 50)
    needed_cols = max(30, len(all_months) + 25)
    ws = _api_retry(
        sh.add_worksheet, title=summary_name,
        rows=needed_rows, cols=needed_cols,
    )
    sheet_id = ws.id

    _api_retry(
        ws.update, range_name="A1",
        values=[["残業時間・特出時間 月別集計 (全スタッフ)"]],
        value_input_option="USER_ENTERED",
    )

    # ===== 4セクション (各セクション: クロステーブル + 全スタッフをまとめた棒グラフ) =====
    # X軸=スタッフ, 系列=月。1グラフに全スタッフ × 全月が並ぶ。
    # 将来 残業申請データが入手できたら、ここに 5番目のセクションとして
    # 「申請 vs 実残業 予実差分」 を追加可能。(下の TODO 参照)

    def _build_table(metric: str, cumulative: bool) -> list[list]:
        """metric: 'over' or 'spec'. cumulative: 月ごとに累積するか。"""
        rows_out: list[list] = [["スタッフ"] + [f"{y}-{m:02d}" for (y, m) in all_months]]
        for staff in all_staff:
            row: list = [staff]
            cum = 0
            for (y, m) in all_months:
                v = monthly.get((staff, y, m), {}).get(metric, 0)
                if cumulative:
                    cum += v
                    row.append(cum / 86400.0)
                else:
                    row.append(v / 86400.0)
            rows_out.append(row)
        return rows_out

    sections = [
        ("残業時間 (月別合計)", _build_table("over", False)),
        ("特出時間 (月別合計)", _build_table("spec", False)),
        ("累積残業時間", _build_table("over", True)),
        ("累積特出時間", _build_table("spec", True)),
    ]

    current_row = 3
    chart_requests = []
    last_col_letter = _col_letter(len(all_months))  # 月列の最右

    for section_title, table in sections:
      try:
        # セクションタイトル
        _api_retry(
            ws.update, range_name=f"A{current_row}",
            values=[[section_title]], value_input_option="USER_ENTERED",
        )
        title_row = current_row
        current_row += 1

        # テーブル (ヘッダ含む) を一括書き込み (BATCH = 200行ずつ)
        table_start = current_row
        BATCH = 200
        for i in range(0, len(table), BATCH):
            chunk = table[i:i + BATCH]
            _api_retry(
                ws.update, range_name=f"A{table_start + i}",
                values=chunk, value_input_option="USER_ENTERED",
            )
        table_end = table_start + len(table) - 1
        logger.info("集計セクション '%s': %d行書き込み (A%d〜%d)",
                    section_title, len(table), table_start, table_end)

        # 月列を時刻フォーマット (データ部分のみ、ヘッダ除く)
        try:
            _api_retry(
                ws.format,
                f"B{table_start + 1}:{last_col_letter}{table_end}",
                {"numberFormat": {"type": "TIME", "pattern": "[h]:mm"}},
            )
        except Exception as e:
            logger.warning("セクション '%s' 時刻フォーマット失敗: %s", section_title, e)

        # 棒グラフ: X=スタッフ, 系列=月 (各月が別色)
        chart_requests.append({
            "addChart": {
                "chart": {
                    "spec": {
                        "title": section_title,
                        "basicChart": {
                            "chartType": "COLUMN",
                            "legendPosition": "BOTTOM_LEGEND",
                            "axis": [
                                {"position": "BOTTOM_AXIS", "title": "スタッフ"},
                                {"position": "LEFT_AXIS", "title": "時間"},
                            ],
                            "domains": [{
                                "domain": {"sourceRange": {"sources": [{
                                    "sheetId": sheet_id,
                                    "startRowIndex": table_start - 1,  # ヘッダ含む
                                    "endRowIndex": table_end,
                                    "startColumnIndex": 0,
                                    "endColumnIndex": 1,
                                }]}}
                            }],
                            "series": [
                                {"series": {"sourceRange": {"sources": [{
                                    "sheetId": sheet_id,
                                    "startRowIndex": table_start - 1,
                                    "endRowIndex": table_end,
                                    "startColumnIndex": ci,
                                    "endColumnIndex": ci + 1,
                                }]}}, "targetAxis": "LEFT_AXIS"}
                                for ci in range(1, len(all_months) + 1)
                            ],
                            "headerCount": 1,
                        }
                    },
                    "position": {
                        "overlayPosition": {
                            "anchorCell": {
                                "sheetId": sheet_id,
                                "rowIndex": title_row - 1,
                                "columnIndex": len(all_months) + 3,
                            },
                            "widthPixels": 1500,
                            "heightPixels": 500,
                        }
                    }
                }
            }
        })

        # 次のセクション開始位置 (テーブル末・チャート末の下、3行空ける)
        chart_height_rows = 27  # 500px ≈ 27行
        current_row = max(table_end + 1, title_row + chart_height_rows) + 3
        time.sleep(0.5)
      except Exception as e:
          logger.error("集計セクション '%s' 失敗 (続行): %s", section_title, e, exc_info=True)
          current_row += 30
          continue

    # ===== 将来拡張ポイント: 残業申請 予実管理 =====
    # 残業申請システムが用意できたとき、ここに5番目のセクションを追加する想定。
    # 必要な作業:
    #   1. 申請データ(スタッフ × 日付 × 申請残業時間)を別シート or API で読み取る関数を追加
    #   2. 月別に集計して `_build_table` 同様のクロステーブルを作る
    #   3. 同じ要領で chart_requests に「申請 vs 実残業 差分」のグラフを append
    # 既存の 4 セクションと同じ枠組みで拡張できるよう、関数とループ構造はそのまま使えます。
    try:
        _api_retry(
            ws.update, range_name=f"A{current_row}",
            values=[["残業申請 予実管理 (実装予定)"]],
            value_input_option="USER_ENTERED",
        )
        current_row += 1
        _api_retry(
            ws.update, range_name=f"A{current_row}",
            values=[[
                "残業申請システム連携後、申請時間 vs 実残業時間の予実比較表とグラフをここに表示します。"
            ]],
            value_input_option="USER_ENTERED",
        )
    except Exception as e:
        logger.warning("予実管理プレースホルダ書き込み失敗: %s", e)

    # チャート一括作成
    if chart_requests:
        try:
            _api_retry(sh.batch_update, {"requests": chart_requests})
            logger.info("集計タブにグラフ %d 個作成", len(chart_requests))
        except Exception as e:
            logger.error("チャート作成失敗: %s", e, exc_info=True)

    logger.info("集計タブ作成完了: 全%dスタッフ × %dヶ月", len(all_staff), len(all_months))


def _load_attendance_sources(file_path: Path) -> list[tuple[str, list[list]]]:
    """ZIPまたはXLSXを読み、(スタッフラベル, 行配列) のリストを返す。

    スタッフラベルはxlsx内部のシート名(openpyxlが正しく復号)を優先。
    シート名でうまく取れない場合は、xlsx内のスタッフ情報行(行3)から
    'スタッフ名(スタッフコード)' を組み立てる。
    """
    import io
    import zipfile
    import openpyxl

    suffix = file_path.suffix.lower()
    result: list[tuple[str, list[list]]] = []

    def _label_from_workbook(wb, ws_name: str, rows: list[list]) -> str:
        # まずシート名を優先 (例: "廣海 愛美")
        label = (ws_name or "").strip()
        # スタッフコードを行3から拾えれば付与
        try:
            r3 = rows[3] if len(rows) > 3 else []
            # 形: ['廣海 愛美', None, '66', None, '医療法人...', ...]
            staff_name = str(r3[0]).strip() if r3 and r3[0] else ""
            staff_code = str(r3[2]).strip() if len(r3) > 2 and r3[2] else ""
            if staff_name:
                # シート名が違えばスタッフ名を優先採用
                if not label or label != staff_name:
                    label = staff_name
                if staff_code and f"({staff_code})" not in label:
                    label = f"{label} ({staff_code})"
        except Exception:
            pass
        return label or "(unknown)"

    if suffix == ".zip":
        with zipfile.ZipFile(file_path) as z:
            names = sorted(n for n in z.namelist() if n.lower().endswith((".xlsx", ".xlsm")))
            for name in names:
                try:
                    data = z.read(name)
                    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
                except Exception as e:
                    logger.warning("ZIP内 %s 読み込み失敗: %s", name, e)
                    continue
                # 通常スタッフ別 xlsx は 1 シート構成
                for ws_name in wb.sheetnames:
                    rows = [list(r) for r in wb[ws_name].iter_rows(values_only=True)]
                    label = _label_from_workbook(wb, ws_name, rows)
                    result.append((label, rows))
    elif suffix in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for ws_name in wb.sheetnames:
            rows = [list(r) for r in wb[ws_name].iter_rows(values_only=True)]
            label = _label_from_workbook(wb, ws_name, rows)
            result.append((label, rows))
    else:
        raise ValueError(f"未対応のファイル形式: {file_path}")

    return result


def _find_header_row(rows: list[list]) -> int | None:
    """各行を見て '日付' を含むセルがあれば、その行をヘッダとみなす。"""
    for i, r in enumerate(rows[:30]):  # 先頭30行以内に必ずある想定
        for c in r:
            if c is not None and "日付" in str(c):
                return i
    return None


def _norm(v) -> str:
    """セル値を文字列化。None→空文字。"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return str(v)


if __name__ == "__main__":
    import json
    import sys

    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python sheets_upload.py <csv_path>            # 勤務データCSV追記")
        print("  python sheets_upload.py --book <xlsx_path>    # 出勤簿Excel追記")
        sys.exit(1)
    cfg = json.loads(Path("config.json").read_text(encoding="utf-8"))
    if sys.argv[1] == "--book":
        n = upload_attendance_book(cfg, Path(sys.argv[2]))
        print(f"出勤簿追記完了: {n}行")
    else:
        n = upload_csv(cfg, Path(sys.argv[1]))
        print(f"勤務データ追記完了: {n}行")
